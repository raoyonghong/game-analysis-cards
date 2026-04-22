import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = r"C:\Users\yonghong.rao\Desktop\Claude_Code_Workspace\02_Projects\加拿大-卡尔加里.xlsx"

# ===================== DATA =====================

TRIP_DATA = [
    {
        "date": "4月27日 (周日)",
        "theme": "抵达卡尔加里 · 市区探索",
        "itinerary": [
            ("15:15", "卡尔加里国际机场 (YYC) 降落", "-", "-"),
            ("15:15~16:30", "入境 / 取行李 / 租车", "75分钟", "-"),
            ("16:30~17:15", "驾车前往酒店办理入住 (市中心)", "45分钟", "约18km"),
            ("17:30~19:00", "卡尔加里塔 (Calgary Tower)", "90分钟", "约2km"),
            ("19:00~20:30", "王子岛公园散步 & 晚餐", "90分钟", "约1.5km"),
            ("20:30", "返回酒店休息", "-", "-"),
        ],
        "attractions": [
            ("卡尔加里塔", "Calgary Tower，俯瞰市区全景，旋转餐厅"),
            ("王子岛公园", "Prince's Island Park，市中心绿洲，适合散步"),
            ("步行街 Stephen Ave", "百年历史商业街，咖啡馆、餐厅林立"),
        ],
        "hotel": "卡尔加里市中心 (推荐: Hyatt Regency Calgary / Hotel Arts)",
        "food": [
            ("Alberta牛排", "阿尔伯塔省以顶级牛肉闻名，必试ribeye/NY strip"),
            ("Poutine", "加拿大国民小吃，炸薯条+芝士+肉汁"),
            ("Calgary Stampede风味烤肉", "牛仔风味BBQ，当地特色"),
            ("Tim Hortons咖啡", "加拿大标志性咖啡连锁，Timbits甜甜圈球"),
        ],
    },
    {
        "date": "4月28日 (周一)",
        "theme": "班夫国家公园 · 硫磺山 · 弓河瀑布",
        "itinerary": [
            ("09:00", "酒店出发前往班夫", "-", "-"),
            ("10:30~11:00", "抵达班夫小镇", "90分钟", "约128km"),
            ("11:00~12:30", "硫磺山缆车 (Sulphur Mountain Gondola)", "90分钟", "约3km"),
            ("12:30~13:30", "山顶徒步 & 远眺班夫小镇", "60分钟", "1.5km徒步"),
            ("13:30~14:30", "午餐 (班夫小镇)", "60分钟", "约3km"),
            ("14:30~16:00", "班夫温泉泳池 (Banff Upper Hot Springs)", "90分钟", "约1km"),
            ("16:00~17:30", "弓河瀑布 (Bow Falls)", "90分钟", "约2km"),
            ("17:30~18:30", "班夫大道购物/闲逛", "60分钟", "步行"),
            ("19:00~20:30", "晚餐后返回卡尔加里", "90分钟", "约128km"),
        ],
        "attractions": [
            ("硫磺山", "Sulphur Mountain，缆车直达山顶，360°雪山全景"),
            ("弓河瀑布", "Bow Falls，班夫地标，电影《史密斯镇》取景地"),
            ("班夫大道", "Banff Avenue，纪念品商店 & 山地装备店"),
            ("班夫温泉", "Banff Upper Hot Springs，历史温泉，雪山环绕泡温泉"),
        ],
        "hotel": "卡尔加里市中心 (同前)",
        "food": [
            ("班夫野牛汉堡", "Bison Burger，当地特色野牛肉汉堡"),
            ("枫糖浆华夫饼", "加拿大枫糖浆，早餐必试"),
            ("鲑鱼配野菜", "洛矶山脉风味餐厅特色菜"),
            ("热可可/苹果派", "山区咖啡馆暖身饮品"),
        ],
    },
    {
        "date": "4月29日 (周二)",
        "theme": "路易斯湖 · 梦莲湖 · 冰原大道",
        "itinerary": [
            ("09:00", "酒店出发前往路易斯湖", "-", "-"),
            ("11:00~12:30", "路易斯湖 (Lake Louise)", "90分钟", "约184km"),
            ("12:30~13:30", "路易斯湖冰川酒店外观 & 湖边徒步", "60分钟", "约3km"),
            ("13:30~15:00", "梦莲湖 (Moraine Lake) - 十峰谷倒影", "90分钟", "约14km"),
            ("15:00~16:00", "路边野餐/简餐", "60分钟", "-"),
            ("16:00~17:30", "冰原大道沿途观景台停留", "90分钟", "约80km单程"),
            ("17:30~18:30", "Peyto Lake 观景台", "60分钟", "步行约2km"),
            ("19:00~20:30", "返回班夫晚餐", "90分钟", "约185km"),
        ],
        "attractions": [
            ("路易斯湖", "Lake Louise，翡翠绿湖+维多利亚冰川，世界最美湖泊之一"),
            ("梦莲湖", "Moraine Lake，十峰谷倒影，加元$20钞票背面图案"),
            ("冰原大道", "Icefields Parkway，全球十大最美公路之一"),
            ("Peyto Lake", "鸢尾色湖水，从观景台俯瞰形如狐狸头"),
        ],
        "hotel": "班夫镇内 (推荐: Fairmont Banff Springs / Rimrock Resort)",
        "food": [
            ("费尔蒙餐厅", "Fairmont Banff Springs酒店内精致料理"),
            ("Balkan Restaurant", "班夫希腊餐厅，当地老字号"),
            ("Wild Flour Bakery", "班夫本地面包咖啡馆，早餐/午餐"),
            ("鹿肉/驼鹿肉料理", "洛矶山脉特色野味料理"),
        ],
    },
    {
        "date": "4月30日 (周三)",
        "theme": "哥伦比亚冰原 · 冰川步道 · 贾斯珀",
        "itinerary": [
            ("09:00", "班夫出发北上", "-", "-"),
            ("11:00~12:30", "哥伦比亚冰原游客中心", "90分钟", "约230km"),
            ("12:30~14:30", "冰川步道 (Glacier Skywalk) 玻璃观景台", "120分钟", "步行约1km"),
            ("14:30~16:00", "冰原雪地车体验 (Brewster Snow Coach)", "90分钟", "冰原上"),
            ("16:00~17:00", "午餐 (游客中心餐厅)", "60分钟", "-"),
            ("17:00~18:30", "继续北行至贾斯珀小镇", "90分钟", "约103km"),
            ("18:30~20:00", "贾斯珀小镇漫步 & 晚餐", "90分钟", "步行"),
            ("20:00", "入住贾斯珀酒店", "-", "-"),
        ],
        "attractions": [
            ("哥伦比亚冰原", "Columbia Icefield，北美最大冰原之一，约325平方公里"),
            ("冰川步道玻璃台", "Glacier Skywalk，悬空280米玻璃地板观景台"),
            ("雪地车体验", "Snow Coach驶上阿萨巴斯卡冰川，可直接触摸万年冰"),
            ("贾斯珀小镇", "Jasper Town，贾斯珀国家公园门户，悠闲山地小镇"),
        ],
        "hotel": "贾斯珀镇内 (推荐: Fairmont Jasper Park Lodge / Jasper Inn)",
        "food": [
            ("冰原游客中心餐厅", "景区内自助餐，暖身热汤必试"),
            ("贾斯珀牛肉派", "Jasper当地牛肉馅饼，山地风味"),
            ("Treeline餐厅", "贾斯珀精致料理，当地食材"),
            ("热苹果酒", "Mulled Apple Cider，山地小镇特色热饮"),
        ],
    },
    {
        "date": "5月1日 (周四)",
        "theme": "马林湖 · 马林峡谷 · 贾斯珀国家公园",
        "itinerary": [
            ("09:00", "贾斯珀酒店出发", "-", "-"),
            ("09:30~11:30", "马林湖 (Maligne Lake) 游船 (Spirit Island)", "120分钟", "约48km"),
            ("11:30~12:30", "湖边徒步", "60分钟", "约3km"),
            ("12:30~13:30", "午餐 (Maligne Lake Chalet)", "60分钟", "-"),
            ("13:30~15:00", "马林峡谷 (Maligne Canyon) 冰桥/峡谷栈道", "90分钟", "约50km"),
            ("15:00~16:30", "Medicine Lake (药湖) 观景", "90分钟", "约22km"),
            ("16:30~18:00", "贾斯珀小镇购物 & 咖啡", "90分钟", "约24km"),
            ("18:00~19:00", "Jasper Planetarium (暗夜星空小镇)", "60分钟", "步行"),
            ("19:30~21:00", "晚餐", "90分钟", "步行"),
        ],
        "attractions": [
            ("马林湖", "Maligne Lake，加拿大洛矶山脉最大天然湖，Spirit Island游船"),
            ("马林峡谷", "Maligne Canyon，深达55米的峡谷，冰桥/栈道徒步"),
            ("药湖", "Medicine Lake，世界罕见的消失湖，秋季湖水会缩减"),
            ("贾斯珀暗夜", "全球最大暗夜公园之一，星空观测绝佳"),
        ],
        "hotel": "贾斯珀镇内 (同前)",
        "food": [
            ("马林湖船屋餐厅", "Maligne Lake Chalet，湖景午餐，三文鱼沙拉"),
            ("Earls Kitchen贾斯珀", "加拿大连锁精致休闲餐厅"),
            ("Syrahs of Jasper", "贾斯珀高档餐厅，洛矶山野味料理"),
            ("熊掌饼干", "Bear Paw Bakery，贾斯珀最有名的肉桂卷&熊掌饼"),
        ],
    },
    {
        "date": "5月2日 (周五)",
        "theme": "返程卡尔加里 · 德拉姆黑勒恐龙谷",
        "itinerary": [
            ("09:00", "贾斯珀出发南下", "-", "-"),
            ("11:00~12:00", "埃德森小镇加油/休息", "60分钟", "约181km"),
            ("12:00~13:30", "继续驾车", "90分钟", "约150km"),
            ("13:30~14:30", "德拉姆黑勒入口休息/午餐", "60分钟", "约340km"),
            ("14:30~17:00", "德拉姆黑勒恐龙谷 (Drumheller Badlands + 胡道大峡谷Hoodoos)", "150分钟", "约3km徒步"),
            ("17:00~18:00", "皇家泰瑞尔博物馆 (Royal Tyrrell Museum)", "60分钟", "约6km"),
            ("18:00~19:30", "驾车返回卡尔加里", "90分钟", "约138km"),
            ("19:30~21:00", "卡尔加里市区庆祝晚餐", "90分钟", "-"),
            ("21:00", "入住机场附近酒店 (方便次日还车)", "-", "-"),
        ],
        "attractions": [
            ("德拉姆黑勒恐龙谷", "Drumheller Badlands，Alberta省独特荒漠地貌，化石圣地"),
            ("胡道奇岩", "Hoodoos，千年风化形成的蘑菇状岩柱，超现实景观"),
            ("皇家泰瑞尔博物馆", "Royal Tyrrell Museum，全球最大恐龙博物馆之一"),
            ("马蹄铁峡谷", "Horseshoe Canyon，色彩斑斓的侵蚀峡谷"),
        ],
        "hotel": "卡尔加里机场附近 (推荐: Delta Hotels Calgary Airport / Holiday Inn Airport)",
        "food": [
            ("德拉姆黑勒肉桂卷", "Drumheller世界最大肉桂卷，当地必打卡"),
            ("Badlands牛肉汉堡", "恐龙谷风格汉堡，荒野风味"),
            ("卡尔加里最后晚餐", "River Cafe或The Nash，告别Alberta美食之夜"),
            ("Alberta啤酒", "Big Rock Brewery，卡尔加里本地精酿啤酒"),
        ],
    },
    {
        "date": "5月3日 (周六)",
        "theme": "卡尔加里市区购物 · 出发回家",
        "itinerary": [
            ("08:30", "酒店退房/吃早餐", "-", "-"),
            ("09:00~10:30", "还车前最后购物 (CF Market Mall / Chinook Centre)", "90分钟", "约15km"),
            ("10:30~11:00", "驾车前往机场还车", "30分钟", "约20km"),
            ("11:00~12:00", "还车 + 机场Check-in办理", "60分钟", "-"),
            ("12:00~14:20", "安检 / 候机 / 机场午餐", "140分钟", "-"),
            ("14:20", "卡尔加里国际机场 (YYC) 起飞", "-", "-"),
        ],
        "attractions": [
            ("CF Market Mall", "卡尔加里大型购物中心，加拿大品牌&纪念品"),
            ("Chinook Centre", "高端购物中心，Coach/LV/Lululemon等品牌"),
            ("YYC机场免税店", "最后购物机会，枫糖浆/冰酒/冰球周边"),
            ("Calgary Stampede纪念品", "牛仔节主题纪念品，当地特色伴手礼"),
        ],
        "hotel": "无 (当日飞离)",
        "food": [
            ("机场早餐/午餐", "YYC机场有Tim Hortons/A&W等连锁"),
            ("枫糖浆系列产品", "带回家的最佳伴手礼：枫糖浆/枫糖饼干"),
            ("加拿大冰酒", "Icewine，全球最佳产地之一，送礼佳品"),
            ("熊掌软糖/鲑鱼零食", "机场免税区特色零食，路上解馋"),
        ],
    },
]

# ===================== STYLES =====================

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL   = hex_fill("1F4E79")   # dark navy
DAY_FILL      = hex_fill("2E75B6")   # medium blue
SECTION_FILLS = {
    "itinerary":   hex_fill("D6E4F0"),
    "attractions": hex_fill("E2EFDA"),
    "hotel":       hex_fill("FFF2CC"),
    "food":        hex_fill("FCE4D6"),
}
SECTION_HEADER_FILLS = {
    "itinerary":   hex_fill("2E75B6"),
    "attractions": hex_fill("548235"),
    "hotel":       hex_fill("C08000"),
    "food":        hex_fill("C55A11"),
}
ROW_ALT = {
    "itinerary":   hex_fill("EBF3FB"),
    "attractions": hex_fill("F2F8ED"),
    "hotel":       hex_fill("FFFBE6"),
    "food":        hex_fill("FDF0E8"),
}

WHITE_FONT  = Font(name="微软雅黑", color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(name="微软雅黑", color="FFFFFF", bold=True, size=14)
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT   = Font(name="微软雅黑", bold=True, size=10)
SECTION_HEADER_FONT = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ===================== BUILD =====================

wb = openpyxl.Workbook()

# ---- Overview sheet ----
ws_ov = wb.active
ws_ov.title = "行程总览"

ws_ov.column_dimensions["A"].width = 18
ws_ov.column_dimensions["B"].width = 30
ws_ov.column_dimensions["C"].width = 45
ws_ov.column_dimensions["D"].width = 28
ws_ov.column_dimensions["E"].width = 28

# Title row
ws_ov.merge_cells("A1:E1")
ws_ov["A1"] = "加拿大卡尔加里 · 洛矶山脉 7日自驾旅行计划"
ws_ov["A1"].fill = HEADER_FILL
ws_ov["A1"].font = TITLE_FONT
ws_ov["A1"].alignment = CENTER
ws_ov.row_dimensions[1].height = 36

# Sub-title info
ws_ov.merge_cells("A2:E2")
ws_ov["A2"] = "出发: 2025年4月27日 15:15 抵达YYC  |  返程: 2025年5月3日 14:20 离开YYC  |  人数: 3人  |  交通: 自驾油车"
ws_ov["A2"].fill = hex_fill("2E75B6")
ws_ov["A2"].font = Font(name="微软雅黑", color="FFFFFF", size=10)
ws_ov["A2"].alignment = CENTER
ws_ov.row_dimensions[2].height = 22

# Column headers
headers = ["日期", "主题", "主要景点", "住宿", "特色美食"]
for c, h in enumerate(headers, 1):
    cell = ws_ov.cell(row=3, column=c, value=h)
    cell.fill = hex_fill("1A3A5C")
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = thin_border()
ws_ov.row_dimensions[3].height = 22

# Data rows
for r, day in enumerate(TRIP_DATA, 4):
    attractions_txt = "\n".join(f"• {a[0]}" for a in day["attractions"])
    hotel_txt = day["hotel"]
    food_txt  = "\n".join(f"• {f[0]}: {f[1]}" for f in day["food"])

    fill = DAY_FILL if r % 2 == 0 else hex_fill("3A86C8")
    row_data = [day["date"], day["theme"], attractions_txt, hotel_txt, food_txt]
    for c, val in enumerate(row_data, 1):
        cell = ws_ov.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = Font(name="微软雅黑", color="FFFFFF", size=9)
        cell.alignment = LEFT if c > 1 else CENTER
        cell.border = thin_border()
    ws_ov.row_dimensions[r].height = 72

# ---- Per-day sheets ----
for day in TRIP_DATA:
    short = day["date"].split(" ")[0]   # e.g. 4月27日
    ws = wb.create_sheet(title=short)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    # Day title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{day['date']}  {day['theme']}"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    current_row = 2

    def write_section_header(ws, row, title, fill):
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].fill = fill
        ws[f"A{row}"].font = SECTION_HEADER_FONT
        ws[f"A{row}"].alignment = CENTER
        ws.row_dimensions[row].height = 20
        return row + 1

    # ---- 1. Daily Itinerary ----
    current_row = write_section_header(ws, current_row, "一、当日行程", SECTION_HEADER_FILLS["itinerary"])

    col_headers = ["时间", "行程事项", "耗时", "公里数"]
    ch_fills = [hex_fill("5B9BD5")] * 4
    for c, (h, f) in enumerate(zip(col_headers, ch_fills), 1):
        cell = ws.cell(row=current_row, column=c, value=h)
        cell.fill = f
        cell.font = Font(name="微软雅黑", color="FFFFFF", bold=True, size=9)
        cell.alignment = CENTER
        cell.border = thin_border()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    for i, (time_, item, dur, km) in enumerate(day["itinerary"]):
        fill = SECTION_FILLS["itinerary"] if i % 2 == 0 else ROW_ALT["itinerary"]
        for c, val in enumerate([time_, item, dur, km], 1):
            cell = ws.cell(row=current_row, column=c, value=val)
            cell.fill = fill
            cell.font = NORMAL_FONT
            cell.alignment = CENTER if c != 2 else LEFT
            cell.border = thin_border()
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    current_row += 1  # spacer

    # ---- 2. Attractions ----
    current_row = write_section_header(ws, current_row, "二、当日主要景点", SECTION_HEADER_FILLS["attractions"])

    col_headers2 = ["景点名称", "景点介绍", "", ""]
    for c, h in enumerate(col_headers2, 1):
        cell = ws.cell(row=current_row, column=c, value=h if c <= 2 else "")
        cell.fill = hex_fill("70AD47")
        cell.font = Font(name="微软雅黑", color="FFFFFF", bold=True, size=9)
        cell.alignment = CENTER
        cell.border = thin_border()
    ws.merge_cells(f"B{current_row}:D{current_row}")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    for i, (name, desc) in enumerate(day["attractions"]):
        fill = SECTION_FILLS["attractions"] if i % 2 == 0 else ROW_ALT["attractions"]
        cell_a = ws.cell(row=current_row, column=1, value=name)
        cell_a.fill = fill; cell_a.font = BOLD_FONT; cell_a.alignment = CENTER; cell_a.border = thin_border()
        cell_b = ws.cell(row=current_row, column=2, value=desc)
        cell_b.fill = fill; cell_b.font = NORMAL_FONT; cell_b.alignment = LEFT; cell_b.border = thin_border()
        ws.merge_cells(f"B{current_row}:D{current_row}")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

    current_row += 1

    # ---- 3. Hotel ----
    current_row = write_section_header(ws, current_row, "三、当日住宿", SECTION_HEADER_FILLS["hotel"])
    ws.merge_cells(f"A{current_row}:D{current_row}")
    ws[f"A{current_row}"] = day["hotel"]
    ws[f"A{current_row}"].fill = SECTION_FILLS["hotel"]
    ws[f"A{current_row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"A{current_row}"].alignment = LEFT
    ws[f"A{current_row}"].border = thin_border()
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    # ---- 4. Food ----
    current_row = write_section_header(ws, current_row, "四、当日特色美食", SECTION_HEADER_FILLS["food"])

    col_headers3 = ["美食名称", "特色描述", "", ""]
    for c, h in enumerate(col_headers3, 1):
        cell = ws.cell(row=current_row, column=c, value=h if c <= 2 else "")
        cell.fill = hex_fill("ED7D31")
        cell.font = Font(name="微软雅黑", color="FFFFFF", bold=True, size=9)
        cell.alignment = CENTER
        cell.border = thin_border()
    ws.merge_cells(f"B{current_row}:D{current_row}")
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    for i, (fname, fdesc) in enumerate(day["food"]):
        fill = SECTION_FILLS["food"] if i % 2 == 0 else ROW_ALT["food"]
        cell_a = ws.cell(row=current_row, column=1, value=fname)
        cell_a.fill = fill; cell_a.font = BOLD_FONT; cell_a.alignment = CENTER; cell_a.border = thin_border()
        cell_b = ws.cell(row=current_row, column=2, value=fdesc)
        cell_b.fill = fill; cell_b.font = NORMAL_FONT; cell_b.alignment = LEFT; cell_b.border = thin_border()
        ws.merge_cells(f"B{current_row}:D{current_row}")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

wb.save(OUTPUT_PATH)
print(f"Excel saved: {OUTPUT_PATH}")
